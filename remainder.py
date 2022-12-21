import openpyxl, smtplib, sys


# Open the spreadsheet and get the latest file status.
wb = openpyxl.load_workbook('Copy of LOU.xlsx')
sheet = wb['Sheet3']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

#Check each member's payment status.
dowloadedMember= {}
for r in range(2, sheet.max_row + 1):
      feedback = sheet.cell(row=r, column=lastCol).value
if feedback != 'dowloaded':
         name = sheet.cell(row=r, column=1).value
         email = sheet.cell(row=r, column=2).value
         dowloadedMember[name] = email


# Log in to email account.
smtpObj = smtplib.SMTP('smtp.outlook.com', 993)   #smtp.office365.com
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('jyothipriyanka.madugula@gmail.com', 'Mpriyanka@462')
sys.argv[1]

#log in to outlook
#from mailbox import create_account
#account = create_account('priyanka.madugula@yash.com', 'Bob#priya@462')

# Send out reminder emails.
for name, email in dowloadedMember.items():
    #string interpolation
     
    body = "Subject: %s undownloadedMember.\nDear %s,\nRecords show that you have not send file for %s. Please send the attachment as soon as possible. Thank you!'"
    (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('priyanka.madugula@yash.com', email,
body)

    if sendmailStatus != {}:
           print('There was a problem sending email to %s: %s' % (email,
           sendmailStatus))
smtpObj.quit()

